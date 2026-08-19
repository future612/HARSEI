#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Render section 3.8 summary figures from revised GeoDetector and RF-SHAP outputs."""

from __future__ import annotations

import csv
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None


ROOT = Path(r"D:\Codex\260724 小论文")
YEARS = [2000, 2005, 2010, 2015, 2020, 2024]
RF_ROOT = ROOT / "意见修改" / "图表" / "HARSEI_RF_SHAP_变量重要性"
RF_TABLE = RF_ROOT / "tables"
GD_TABLE = ROOT / "意见修改" / "图表" / "HARSEI_驱动因子交互探测" / "tables"
FIG_DIR = RF_ROOT / "figures"
TABLE_DIR = RF_ROOT / "tables"

PALETTE = {
    "blue": "#376795",
    "teal": "#2a9d8f",
    "green": "#59a14f",
    "gold": "#e9c46a",
    "coral": "#e76f51",
    "navy": "#22304a",
    "gray": "#6c757d",
}


def read_metrics() -> List[Dict[str, float]]:
    rows = []
    with (RF_TABLE / "RF_SHAP_model_metrics.csv").open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rows.append(
                {
                    "Year": int(row["Year"]),
                    "N_valid_pixels": int(row["N_valid_pixels"]),
                    "N_model_samples": int(row["N_model_samples"]),
                    "N_train": int(row["N_train"]),
                    "N_test": int(row["N_test"]),
                    "N_SHAP": int(row["N_SHAP"]),
                    "RMSE": float(row["RMSE"]),
                    "R2": float(row["R2"]),
                }
            )
    return rows


def read_rf_importance() -> Dict[int, Dict[str, float]]:
    vals: Dict[int, Dict[str, float]] = defaultdict(dict)
    with (RF_TABLE / "RF_SHAP_variable_importance.csv").open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            vals[int(row["Year"])][row["Variable"]] = float(row["Mean_abs_SHAP"])
    return vals


def read_gd_q() -> Dict[int, Dict[str, float]]:
    vals: Dict[int, Dict[str, float]] = defaultdict(dict)
    for year in YEARS:
        path = GD_TABLE / f"GD_factor_revised_HARSEI_{year}.csv"
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                vals[year][row["Factor"]] = float(row["q"])
    return vals


def relative_by_year(values: Dict[int, Dict[str, float]], variables: List[str]) -> Dict[int, Dict[str, float]]:
    out: Dict[int, Dict[str, float]] = defaultdict(dict)
    for year in YEARS:
        total = sum(values[year].get(v, 0.0) for v in variables)
        for var in variables:
            out[year][var] = values[year].get(var, 0.0) / (total + 1e-12)
    return out


def save_fig(fig, stem: str) -> None:
    for ext, dpi in [("png", 450), ("tif", 450), ("pdf", 450)]:
        path = FIG_DIR / f"{stem}.{ext}"
        if ext == "tif":
            fig.savefig(path, dpi=dpi, pil_kwargs={"compression": "tiff_lzw"})
        else:
            fig.savefig(path, dpi=dpi)


def figure13(gd: Dict[int, Dict[str, float]], rf: Dict[int, Dict[str, float]], variables: List[str]) -> List[Dict[str, float]]:
    rf_rel = relative_by_year(rf, variables)
    rows = []
    for var in variables:
        gd_mean = statistics.mean(gd[y][var] for y in YEARS)
        rf_mean = statistics.mean(rf_rel[y][var] for y in YEARS)
        rows.append({"Variable": var, "GeoDetector_mean_q": gd_mean, "RF_SHAP_mean_relative": rf_mean})
    rows.sort(key=lambda r: r["GeoDetector_mean_q"], reverse=True)

    with (TABLE_DIR / "Fig13_exogenous_driver_importance_after_excluding_HARSEI_components.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Variable", "GeoDetector_mean_q", "RF_SHAP_mean_relative"])
        writer.writeheader()
        writer.writerows(rows)

    gd_order = sorted(rows, key=lambda r: r["GeoDetector_mean_q"])
    rf_order = sorted(rows, key=lambda r: r["RF_SHAP_mean_relative"])

    fig, axes = plt.subplots(1, 2, figsize=(10.8, 5.2), constrained_layout=True)
    for ax, data, key, title, xlabel, color in [
        (axes[0], gd_order, "GeoDetector_mean_q", "(a) GeoDetector", "Mean q statistic", PALETTE["blue"]),
        (axes[1], rf_order, "RF_SHAP_mean_relative", "(b) RF-SHAP", "Mean relative |SHAP|", PALETTE["teal"]),
    ]:
        labels = [r["Variable"] for r in data]
        values = [r[key] for r in data]
        y = np.arange(len(labels))
        ax.barh(y, values, color=color, edgecolor="#243447", linewidth=0.45, alpha=0.92)
        ax.set_yticks(y)
        ax.set_yticklabels(labels, fontsize=9)
        ax.set_xlabel(xlabel)
        ax.set_title(title, loc="left", fontweight="bold")
        ax.grid(axis="x", color="#e8e8e8", linewidth=0.7)
        ax.set_axisbelow(True)
        xmax = max(values) * 1.12
        ax.set_xlim(0, xmax)
        for yi, value in zip(y, values):
            ax.text(value + xmax * 0.012, yi, f"{value:.3f}" if key.startswith("Geo") else f"{value:.2f}", va="center", fontsize=8)
        for spine in ("top", "right"):
            ax.spines[spine].set_visible(False)
    fig.suptitle("Fig. 13. Exogenous driver importance after excluding HARSEI components", fontsize=13.5, fontweight="bold")
    save_fig(fig, "Fig13_exogenous_driver_importance_after_excluding_HARSEI_components")
    plt.close(fig)
    return rows


def figure14(rf: Dict[int, Dict[str, float]], variables: List[str]) -> List[Dict[str, float]]:
    rf_rel = relative_by_year(rf, variables)
    mean_rel = {var: statistics.mean(rf_rel[y][var] for y in YEARS) for var in variables}
    order = sorted(variables, key=lambda v: mean_rel[v], reverse=True)
    mat = np.array([[rf_rel[y][v] for y in YEARS] for v in order], dtype=float)

    rows = []
    for var in order:
        for year in YEARS:
            rows.append({"Variable": var, "Year": year, "Mean_abs_SHAP": rf[year][var], "Relative_mean_abs_SHAP": rf_rel[year][var]})
    with (TABLE_DIR / "Fig14_multiyear_RF_SHAP_exogenous_driver_importance.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Variable", "Year", "Mean_abs_SHAP", "Relative_mean_abs_SHAP"])
        writer.writeheader()
        writer.writerows(rows)

    cmap = LinearSegmentedColormap.from_list("soft_teal_navy", ["#f5fbfa", "#c9e8df", "#7cc8b2", "#359c9c", "#1f5673", "#17233d"], N=256)
    fig, ax = plt.subplots(figsize=(8.2, 6.0), constrained_layout=True)
    im = ax.imshow(mat, cmap=cmap, aspect="auto")
    ax.set_xticks(range(len(YEARS)))
    ax.set_xticklabels([str(y) for y in YEARS], fontsize=9)
    ax.set_yticks(range(len(order)))
    ax.set_yticklabels(order, fontsize=9)
    ax.set_xlabel("Year")
    ax.set_title("Fig. 14. Multi-year RF-SHAP importance of exogenous drivers", fontsize=13, fontweight="bold")
    for i in range(mat.shape[0]):
        for j in range(mat.shape[1]):
            val = mat[i, j]
            color = "white" if val >= 0.14 else PALETTE["navy"]
            ax.text(j, i, f"{val:.2f}", ha="center", va="center", fontsize=8, color=color)
    cb = fig.colorbar(im, ax=ax, fraction=0.042, pad=0.025)
    cb.set_label("Relative mean |SHAP|")
    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
    save_fig(fig, "Fig14_multiyear_RF_SHAP_exogenous_driver_importance")
    plt.close(fig)
    return rows


def figure15(metrics: List[Dict[str, float]]) -> None:
    with (TABLE_DIR / "Fig15_RF_model_performance_target_years.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Year", "N_valid_pixels", "N_model_samples", "N_train", "N_test", "N_SHAP", "RMSE", "R2"],
        )
        writer.writeheader()
        writer.writerows(metrics)

    years = [m["Year"] for m in metrics]
    r2 = np.array([m["R2"] for m in metrics], dtype=float)
    rmse = np.array([m["RMSE"] for m in metrics], dtype=float)

    fig, axes = plt.subplots(1, 2, figsize=(9.8, 4.3), constrained_layout=True)
    for ax, values, title, ylabel, color, ylim_pad in [
        (axes[0], r2, "(a) Coefficient of determination", "R²", PALETTE["green"], 0.006),
        (axes[1], rmse, "(b) Root mean square error", "RMSE", PALETTE["coral"], 0.002),
    ]:
        ax.plot(years, values, color=color, linewidth=2.2, marker="o", markersize=6.5)
        ax.fill_between(years, values, values.mean(), color=color, alpha=0.12)
        ax.axhline(values.mean(), color=PALETTE["gray"], linestyle="--", linewidth=1.0)
        ax.text(years[-1] + 0.15, values.mean(), f"Mean={values.mean():.3f}", va="center", fontsize=8, color=PALETTE["gray"])
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_xlabel("Year")
        ax.set_ylabel(ylabel)
        ax.set_xticks(years)
        ax.grid(color="#e9ecef", linewidth=0.7)
        ax.set_axisbelow(True)
        ax.set_ylim(values.min() - ylim_pad, values.max() + ylim_pad)
        for x, y in zip(years, values):
            ax.text(x, y + (ylim_pad * 0.35 if ylabel == "R²" else ylim_pad * 0.22), f"{y:.3f}", ha="center", va="bottom", fontsize=8)
        for spine in ("top", "right"):
            ax.spines[spine].set_visible(False)
    fig.suptitle("Fig. 15. Random forest model performance for target years", fontsize=13.5, fontweight="bold")
    save_fig(fig, "Fig15_RF_model_performance_target_years")
    plt.close(fig)


def write_workbook(fig13_rows, fig14_rows, metrics) -> None:
    if Workbook is None:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Fig13_importance"
    ws.append(["Variable", "GeoDetector_mean_q", "RF_SHAP_mean_relative"])
    for row in fig13_rows:
        ws.append([row["Variable"], row["GeoDetector_mean_q"], row["RF_SHAP_mean_relative"]])

    ws = wb.create_sheet("Fig14_RF_SHAP")
    ws.append(["Variable", "Year", "Mean_abs_SHAP", "Relative_mean_abs_SHAP"])
    for row in fig14_rows:
        ws.append([row["Variable"], row["Year"], row["Mean_abs_SHAP"], row["Relative_mean_abs_SHAP"]])

    ws = wb.create_sheet("Fig15_RF_metrics")
    ws.append(["Year", "N_valid_pixels", "N_model_samples", "N_train", "N_test", "N_SHAP", "RMSE", "R2"])
    for row in metrics:
        ws.append([row["Year"], row["N_valid_pixels"], row["N_model_samples"], row["N_train"], row["N_test"], row["N_SHAP"], row["RMSE"], row["R2"]])

    wb.save(TABLE_DIR / "Section_3.8_Fig13_Fig14_Fig15_plot_data.xlsx")


def write_paragraph(metrics: List[Dict[str, float]], rf: Dict[int, Dict[str, float]], variables: List[str]) -> None:
    r2 = [m["R2"] for m in metrics]
    rmse = [m["RMSE"] for m in metrics]
    rf_rel = relative_by_year(rf, variables)
    mean_rel = {var: statistics.mean(rf_rel[y][var] for y in YEARS) for var in variables}
    top_vars = sorted(variables, key=lambda v: mean_rel[v], reverse=True)[:4]
    top_text = "、".join(top_vars)
    cn = (
        "随机森林与 SHAP 分析进一步确认了外生水分-土壤控制对 HARSEI 空间分异的重要作用。"
        "每个目标年份均随机抽取 100,000 个有效像元构建随机森林回归模型，其中 75% 用于训练，25% 用于独立验证，并从验证集中抽取 2,000 个样本计算 SHAP 值。"
        f"六个目标年份模型表现稳定，平均 R² = {statistics.mean(r2):.3f}（范围 {min(r2):.3f}–{max(r2):.3f}），"
        f"平均 RMSE = {statistics.mean(rmse):.3f}（范围 {min(rmse):.3f}–{max(rmse):.3f}）。"
        f"变量重要性结果显示，{top_text} 在多数年份位居 SHAP 贡献前列，说明实际蒸散、降水输入、土壤背景及雪水过程共同控制了区域生态质量的空间差异。"
        "SHAP 蜂群图进一步表明，高、低变量取值对 HARSEI 的贡献方向存在明显差异，反映出驱动因子对生态质量的影响具有非线性和阈值响应特征。"
    )
    en = (
        "The random forest and SHAP analyses further confirmed the importance of exogenous water-soil controls in explaining the spatial differentiation of HARSEI. "
        "For each target year, 100,000 valid pixels were randomly sampled to construct the RF regression model, with 75% used for training and 25% for independent validation; SHAP values were calculated using 2,000 validation samples. "
        f"The six target-year models showed stable performance, with a mean R² of {statistics.mean(r2):.3f} (range: {min(r2):.3f}–{max(r2):.3f}) and a mean RMSE of {statistics.mean(rmse):.3f} (range: {min(rmse):.3f}–{max(rmse):.3f}). "
        f"The importance results indicated that {', '.join(top_vars)} generally ranked among the leading SHAP contributors, suggesting that actual evapotranspiration, precipitation input, soil background, and snow-water processes jointly shaped the spatial variability of regional ecological quality. "
        "The SHAP beeswarm plots further showed that high and low feature values had distinct directions of contribution to HARSEI, implying nonlinear and threshold-dependent ecological responses to the driving factors."
    )
    (RF_ROOT / "Section_3.8_RF_SHAP_revised_paragraph_CN_EN.txt").write_text(
        "【中文正文段落】\n" + cn + "\n\n【English manuscript-ready paragraph】\n" + en + "\n",
        encoding="utf-8-sig",
    )


def main() -> None:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    metrics = read_metrics()
    rf = read_rf_importance()
    gd = read_gd_q()
    variables = sorted(set(rf[YEARS[0]].keys()) & set(gd[YEARS[0]].keys()))
    fig13_rows = figure13(gd, rf, variables)
    fig14_rows = figure14(rf, variables)
    figure15(metrics)
    write_workbook(fig13_rows, fig14_rows, metrics)
    write_paragraph(metrics, rf, variables)
    print(f"Saved section 3.8 figures to: {FIG_DIR}")
    print(f"Saved section 3.8 tables to: {TABLE_DIR}")


if __name__ == "__main__":
    main()
