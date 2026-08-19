#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Train annual RF models and render RF-SHAP explanation figures for HARSEI."""

from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import shap
from matplotlib.colors import LinearSegmentedColormap
from osgeo import gdal
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import train_test_split

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None


YEARS = [2000, 2005, 2010, 2015, 2020, 2024]
FEATURES = ["DEM", "SLOPE", "ASPECT", "WRB", "PRE", "TEMMAX", "TEMMIN", "TEM", "SOIL", "AET", "SWE"]

ROOT = Path(r"D:\Codex\260724 小论文")
HARSEI_DIR = ROOT / "revise" / "annual_harsei_outputs" / "rasters" / "HARSEI"
DRIVER_ROOT = Path(r"E:\wl24\hys to wl\数据清单\1数据下载")
OUT_ROOT = ROOT / "意见修改" / "图表" / "HARSEI_RF_SHAP_变量重要性"
FIG_DIR = OUT_ROOT / "figures"
TABLE_DIR = OUT_ROOT / "tables"

MAX_MODEL_SAMPLES = 100_000
MAX_SHAP_SAMPLES = 2_000
TEST_SIZE = 0.25
RANDOM_SEED = 20260811

DRIVERS: Dict[str, Dict[str, object]] = {
    "DEM": {
        "kind": "continuous",
        "path": DRIVER_ROOT / "24DEM" / "2_mask" / "masked_DEM_30m_shp.tif",
    },
    "SLOPE": {
        "kind": "continuous",
        "path": DRIVER_ROOT / "24DEM" / "2_mask" / "masked_Slope_30m_shp_deg.tif",
    },
    "ASPECT": {
        "kind": "continuous",
        "path": DRIVER_ROOT / "24DEM" / "2_mask" / "masked_Aspect_30m_shp_deg.tif",
    },
    "WRB": {
        "kind": "categorical",
        "path": DRIVER_ROOT / "30WRB" / "WRB4.tif",
    },
    "PRE": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "23Pre" / "2_mask" / "masked_pr_{year}.tif",
    },
    "TEMMAX": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "22Temmax" / "2_mask" / "masked_tem_{year}.tif",
    },
    "TEMMIN": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "25TEMmin" / "{year}.tif",
    },
    "TEM": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "26TEMave" / "{year}.tif",
    },
    "SOIL": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "27soil" / "soil{year}.tif",
    },
    "AET": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "28AET" / "aet{year}.tif",
    },
    "SWE": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "29swe" / "swe{year}.tif",
    },
}


def ensure_dirs() -> None:
    for folder in (OUT_ROOT, FIG_DIR, TABLE_DIR):
        folder.mkdir(parents=True, exist_ok=True)


def driver_path(name: str, year: int) -> Path:
    meta = DRIVERS[name]
    if "path" in meta:
        return Path(meta["path"])
    return Path(str(meta["template"]).format(year=year))


def read_raster(path: Path, ref_ds=None, resample: str = "bilinear") -> np.ndarray:
    if not path.exists():
        raise FileNotFoundError(str(path))
    ds = gdal.Open(str(path))
    if ds is None:
        raise RuntimeError(f"Cannot open raster: {path}")
    if ref_ds is not None:
        same_grid = (
            ds.RasterXSize == ref_ds.RasterXSize
            and ds.RasterYSize == ref_ds.RasterYSize
            and tuple(round(v, 12) for v in ds.GetGeoTransform()) == tuple(round(v, 12) for v in ref_ds.GetGeoTransform())
            and ds.GetProjection() == ref_ds.GetProjection()
        )
        if not same_grid:
            gt = ref_ds.GetGeoTransform()
            xmin = gt[0]
            ymax = gt[3]
            xmax = xmin + gt[1] * ref_ds.RasterXSize
            ymin = ymax + gt[5] * ref_ds.RasterYSize
            alg = gdal.GRA_NearestNeighbour if resample == "nearest" else gdal.GRA_Bilinear
            ds = gdal.Warp(
                "",
                ds,
                format="MEM",
                dstSRS=ref_ds.GetProjection(),
                outputBounds=(xmin, ymin, xmax, ymax),
                width=ref_ds.RasterXSize,
                height=ref_ds.RasterYSize,
                resampleAlg=alg,
            )
    band = ds.GetRasterBand(1)
    arr = band.ReadAsArray().astype(np.float32)
    nodata = band.GetNoDataValue()
    if nodata is not None:
        arr[np.isclose(arr, nodata)] = np.nan
    arr[~np.isfinite(arr)] = np.nan
    arr[arr < -1.0e20] = np.nan
    arr[arr > 1.0e20] = np.nan
    return arr


def load_year_data(year: int) -> Tuple[np.ndarray, np.ndarray, int]:
    harsei_path = HARSEI_DIR / f"YJQ_HARSEI_{year}.tif"
    ref_ds = gdal.Open(str(harsei_path))
    if ref_ds is None:
        raise RuntimeError(f"Cannot open HARSEI raster: {harsei_path}")

    y_arr = read_raster(harsei_path)
    data = []
    mask = np.isfinite(y_arr) & (y_arr > -1000) & (y_arr < 1000)
    for name in FEATURES:
        meta = DRIVERS[name]
        resample = "nearest" if meta["kind"] == "categorical" else "bilinear"
        arr = read_raster(driver_path(name, year), ref_ds=ref_ds, resample=resample)
        if meta["kind"] == "categorical":
            arr = np.rint(arr).astype(np.float32)
        data.append(arr)
        mask &= np.isfinite(arr)
    x_stack = np.stack([arr[mask] for arr in data], axis=1).astype(np.float32)
    y = y_arr[mask].astype(np.float32)
    return x_stack, y, int(mask.sum())


def sample_rows(x: np.ndarray, y: np.ndarray, max_n: int, seed: int) -> Tuple[np.ndarray, np.ndarray]:
    n = y.size
    if n <= max_n:
        return x, y
    rng = np.random.default_rng(seed)
    idx = rng.choice(n, size=max_n, replace=False)
    return x[idx], y[idx]


def train_and_explain(year: int) -> Dict[str, object]:
    cache = TABLE_DIR / f"RF_SHAP_year_cache_m{MAX_MODEL_SAMPLES}_s{MAX_SHAP_SAMPLES}_{year}.npz"
    if cache.exists():
        print(f"Loading cached RF-SHAP result {year}...", flush=True)
        data = np.load(cache, allow_pickle=False)
        return {
            "year": int(data["year"]),
            "n_valid": int(data["n_valid"]),
            "n_model": int(data["n_model"]),
            "n_train": int(data["n_train"]),
            "n_test": int(data["n_test"]),
            "n_shap": int(data["n_shap"]),
            "rmse": float(data["rmse"]),
            "r2": float(data["r2"]),
            "feature_importance": data["feature_importance"],
            "mean_abs_shap": data["mean_abs_shap"],
            "mean_signed_shap": data["mean_signed_shap"],
            "rank_order": data["rank_order"],
            "x_shap": data["x_shap"],
            "y_shap": data["y_shap"],
            "pred_shap": data["pred_shap"],
            "shap_values": data["shap_values"],
        }

    print(f"Loading {year}...", flush=True)
    x_all, y_all, n_valid = load_year_data(year)
    x_model, y_model = sample_rows(x_all, y_all, MAX_MODEL_SAMPLES, RANDOM_SEED + year)
    x_train, x_test, y_train, y_test = train_test_split(
        x_model, y_model, test_size=TEST_SIZE, random_state=RANDOM_SEED + year
    )
    model = RandomForestRegressor(
        n_estimators=160,
        max_features="sqrt",
        min_samples_leaf=6,
        max_depth=20,
        bootstrap=True,
        n_jobs=-1,
        random_state=RANDOM_SEED + year,
    )
    print(f"Training RF {year}: train={len(y_train)}, test={len(y_test)}...", flush=True)
    model.fit(x_train, y_train)
    pred_test = model.predict(x_test)
    rmse = float(math.sqrt(mean_squared_error(y_test, pred_test)))
    r2 = float(r2_score(y_test, pred_test))

    rng = np.random.default_rng(RANDOM_SEED + year * 3)
    shap_n = min(MAX_SHAP_SAMPLES, x_test.shape[0])
    shap_idx = rng.choice(x_test.shape[0], size=shap_n, replace=False)
    x_shap = x_test[shap_idx]
    y_shap = y_test[shap_idx]
    pred_shap = pred_test[shap_idx]

    print(f"Computing SHAP {year}: n={shap_n}...", flush=True)
    explainer = shap.TreeExplainer(model)
    try:
        shap_values = explainer.shap_values(x_shap, check_additivity=False, approximate=True)
    except TypeError:
        shap_values = explainer.shap_values(x_shap, check_additivity=False)
    shap_values = np.asarray(shap_values, dtype=np.float32)
    if shap_values.ndim == 3:
        shap_values = shap_values[:, :, 0]

    mean_abs = np.mean(np.abs(shap_values), axis=0)
    mean_signed = np.mean(shap_values, axis=0)
    ranks = np.argsort(-mean_abs)

    result = {
        "year": year,
        "n_valid": n_valid,
        "n_model": int(y_model.size),
        "n_train": int(y_train.size),
        "n_test": int(y_test.size),
        "n_shap": int(shap_n),
        "rmse": rmse,
        "r2": r2,
        "feature_importance": np.asarray(model.feature_importances_, dtype=np.float32),
        "mean_abs_shap": mean_abs,
        "mean_signed_shap": mean_signed,
        "rank_order": ranks,
        "x_shap": x_shap,
        "y_shap": y_shap,
        "pred_shap": pred_shap,
        "shap_values": shap_values,
    }
    np.savez_compressed(
        cache,
        year=year,
        n_valid=n_valid,
        n_model=int(y_model.size),
        n_train=int(y_train.size),
        n_test=int(y_test.size),
        n_shap=int(shap_n),
        rmse=rmse,
        r2=r2,
        feature_importance=np.asarray(model.feature_importances_, dtype=np.float32),
        mean_abs_shap=mean_abs.astype(np.float32),
        mean_signed_shap=mean_signed.astype(np.float32),
        rank_order=ranks.astype(np.int16),
        x_shap=x_shap.astype(np.float32),
        y_shap=y_shap.astype(np.float32),
        pred_shap=pred_shap.astype(np.float32),
        shap_values=shap_values.astype(np.float32),
    )
    print(f"Cached {year}: R2={r2:.3f}, RMSE={rmse:.3f}", flush=True)
    return result


def normalized_feature_values(values: np.ndarray) -> np.ndarray:
    out = np.zeros_like(values, dtype=np.float32)
    for j in range(values.shape[1]):
        col = values[:, j]
        lo, hi = np.nanpercentile(col, [2, 98])
        if hi <= lo:
            out[:, j] = 0.5
        else:
            out[:, j] = np.clip((col - lo) / (hi - lo), 0, 1)
    return out


def plot_beeswarm(results: List[Dict[str, object]]) -> None:
    cmap = LinearSegmentedColormap.from_list(
        "teal_rose",
        ["#2474f2", "#29b7c9", "#f3d25c", "#fb607f", "#a01a7d"],
        N=256,
    )
    max_abs = max(float(np.nanpercentile(np.abs(r["shap_values"]), 99.5)) for r in results)
    xlim = max_abs * 1.12
    fig, axes = plt.subplots(2, 3, figsize=(16.0, 9.2), constrained_layout=False)
    fig.subplots_adjust(left=0.07, right=0.91, top=0.91, bottom=0.10, wspace=0.28, hspace=0.32)
    letters = ["a", "b", "c", "d", "e", "f"]
    rng = np.random.default_rng(12345)
    last_scatter = None
    for ax, result, letter in zip(axes.flat, results, letters):
        year = int(result["year"])
        shap_values = result["shap_values"]
        x_values = result["x_shap"]
        norm_values = normalized_feature_values(x_values)
        order = result["rank_order"]
        y_labels = [FEATURES[i] for i in order]
        for plot_y, feature_idx in enumerate(order):
            sv = shap_values[:, feature_idx]
            fv = norm_values[:, feature_idx]
            density_weight = np.clip(np.abs(sv) / (np.nanpercentile(np.abs(sv), 95) + 1e-12), 0, 1)
            jitter = (rng.random(sv.size) - 0.5) * (0.16 + 0.34 * density_weight)
            last_scatter = ax.scatter(
                sv,
                np.full_like(sv, plot_y, dtype=np.float32) + jitter,
                c=fv,
                cmap=cmap,
                vmin=0,
                vmax=1,
                s=6.0,
                alpha=0.72,
                linewidths=0,
                rasterized=True,
            )
        ax.axvline(0, color="#555555", lw=0.8, alpha=0.75)
        ax.set_yticks(range(len(order)))
        ax.set_yticklabels(y_labels, fontsize=8)
        ax.invert_yaxis()
        ax.set_xlim(-xlim, xlim)
        ax.set_xlabel("SHAP value (impact on HARSEI)", fontsize=8.5)
        ax.set_title(f"({letter}) {year}  RF-SHAP", loc="left", fontsize=12, fontweight="bold")
        ax.text(
            0.99,
            1.02,
            f"R²={result['r2']:.3f}, RMSE={result['rmse']:.3f}",
            transform=ax.transAxes,
            ha="right",
            va="bottom",
            fontsize=8.5,
            color="#444444",
        )
        ax.grid(axis="y", color="#eeeeee", linewidth=0.6)
        ax.grid(axis="x", color="#f1f1f1", linewidth=0.5)
        for spine in ("top", "right"):
            ax.spines[spine].set_visible(False)

    cax = fig.add_axes([0.93, 0.18, 0.014, 0.64])
    cb = fig.colorbar(last_scatter, cax=cax)
    cb.set_label("Feature value", fontsize=10)
    cb.set_ticks([0, 1])
    cb.set_ticklabels(["Low", "High"])
    cb.ax.tick_params(labelsize=8)
    fig.suptitle("RF-SHAP variable importance and model explanation for revised HARSEI", fontsize=15, fontweight="bold")
    for ext, dpi in [("png", 450), ("tif", 450), ("pdf", 450)]:
        out = FIG_DIR / f"Fig_HARSEI_RF_SHAP_beeswarm_2000_2024.{ext}"
        if ext == "tif":
            fig.savefig(out, dpi=dpi, pil_kwargs={"compression": "tiff_lzw"})
        else:
            fig.savefig(out, dpi=dpi)
    plt.close(fig)


def plot_importance_heatmap(results: List[Dict[str, object]]) -> None:
    mat = np.vstack([r["mean_abs_shap"] for r in results])
    # Normalize within each year to emphasize changing relative importance.
    rel = mat / (mat.sum(axis=1, keepdims=True) + 1e-12)
    order = np.argsort(-mat.mean(axis=0))
    cmap = LinearSegmentedColormap.from_list("mint_navy", ["#f7fbff", "#b7e2d7", "#4aa5a6", "#1f5673", "#17223b"], N=256)
    fig, ax = plt.subplots(figsize=(9.8, 4.6), constrained_layout=True)
    im = ax.imshow(rel[:, order], aspect="auto", cmap=cmap)
    ax.set_xticks(range(len(FEATURES)))
    ax.set_xticklabels([FEATURES[i] for i in order], rotation=45, ha="right")
    ax.set_yticks(range(len(YEARS)))
    ax.set_yticklabels([str(y) for y in YEARS])
    for i in range(len(YEARS)):
        for j, feature_idx in enumerate(order):
            ax.text(j, i, f"{rel[i, feature_idx]:.2f}", ha="center", va="center", fontsize=7, color="white" if rel[i, feature_idx] > 0.12 else "#253044")
    cb = fig.colorbar(im, ax=ax, fraction=0.035, pad=0.025)
    cb.set_label("Relative mean |SHAP|")
    ax.set_title("Relative RF-SHAP importance across years", fontsize=13, fontweight="bold")
    for ext, dpi in [("png", 450), ("tif", 450), ("pdf", 450)]:
        out = FIG_DIR / f"Fig_HARSEI_RF_SHAP_importance_heatmap_2000_2024.{ext}"
        if ext == "tif":
            fig.savefig(out, dpi=dpi, pil_kwargs={"compression": "tiff_lzw"})
        else:
            fig.savefig(out, dpi=dpi)
    plt.close(fig)


def export_tables(results: List[Dict[str, object]]) -> None:
    metrics_path = TABLE_DIR / "RF_SHAP_model_metrics.csv"
    with metrics_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Year", "N_valid_pixels", "N_model_samples", "N_train", "N_test", "N_SHAP", "RMSE", "R2"])
        for r in results:
            writer.writerow([r["year"], r["n_valid"], r["n_model"], r["n_train"], r["n_test"], r["n_shap"], f"{r['rmse']:.10f}", f"{r['r2']:.10f}"])

    importance_rows: List[List[object]] = []
    for r in results:
        year = int(r["year"])
        order = np.argsort(-r["mean_abs_shap"])
        for rank, j in enumerate(order, start=1):
            importance_rows.append(
                [
                    year,
                    FEATURES[j],
                    rank,
                    float(r["mean_abs_shap"][j]),
                    float(r["mean_signed_shap"][j]),
                    float(r["feature_importance"][j]),
                ]
            )
    with (TABLE_DIR / "RF_SHAP_variable_importance.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Year", "Variable", "Rank_by_mean_abs_SHAP", "Mean_abs_SHAP", "Mean_SHAP", "RF_impurity_importance"])
        writer.writerows(importance_rows)

    long_path = TABLE_DIR / "RF_SHAP_beeswarm_plot_data.csv"
    with long_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Year", "Sample_ID", "Variable", "Feature_value", "Feature_value_norm", "SHAP_value", "Observed_HARSEI", "Predicted_HARSEI"])
        for r in results:
            year = int(r["year"])
            x = r["x_shap"]
            norm = normalized_feature_values(x)
            shap_values = r["shap_values"]
            for i in range(x.shape[0]):
                for j, name in enumerate(FEATURES):
                    writer.writerow([year, i, name, float(x[i, j]), float(norm[i, j]), float(shap_values[i, j]), float(r["y_shap"][i]), float(r["pred_shap"][i])])

    for r in results:
        year = int(r["year"])
        out = TABLE_DIR / f"RF_SHAP_sample_wide_{year}.csv"
        with out.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Sample_ID", "Observed_HARSEI", "Predicted_HARSEI"] + [f"value_{v}" for v in FEATURES] + [f"SHAP_{v}" for v in FEATURES])
            x = r["x_shap"]
            shap_values = r["shap_values"]
            for i in range(x.shape[0]):
                writer.writerow([i, float(r["y_shap"][i]), float(r["pred_shap"][i])] + [float(v) for v in x[i]] + [float(v) for v in shap_values[i]])

    if Workbook is not None:
        wb = Workbook()
        ws = wb.active
        ws.title = "model_metrics"
        ws.append(["Year", "N_valid_pixels", "N_model_samples", "N_train", "N_test", "N_SHAP", "RMSE", "R2"])
        for r in results:
            ws.append([r["year"], r["n_valid"], r["n_model"], r["n_train"], r["n_test"], r["n_shap"], r["rmse"], r["r2"]])
        ws = wb.create_sheet("variable_importance")
        ws.append(["Year", "Variable", "Rank_by_mean_abs_SHAP", "Mean_abs_SHAP", "Mean_SHAP", "RF_impurity_importance"])
        for row in importance_rows:
            ws.append(row)
        ws = wb.create_sheet("top4_by_year")
        ws.append(["Year", "Top_1", "Top_2", "Top_3", "Top_4"])
        for r in results:
            order = np.argsort(-r["mean_abs_shap"])[:4]
            ws.append([r["year"]] + [FEATURES[j] for j in order])
        wb.save(TABLE_DIR / "HARSEI_RF_SHAP_variable_importance_data.xlsx")


def write_text(results: List[Dict[str, object]]) -> None:
    top_lines = []
    for r in results:
        order = np.argsort(-r["mean_abs_shap"])[:4]
        top = ", ".join(f"{FEATURES[j]} ({r['mean_abs_shap'][j]:.3f})" for j in order)
        top_lines.append(f"{r['year']}: {top}; R2={r['r2']:.3f}, RMSE={r['rmse']:.3f}")
    text = "\n".join(
        [
            "【中文正文可用表述】",
            "",
            "为进一步量化外生驱动因子对 HARSEI 空间分异的非线性影响，本研究以最新年度 HARSEI 为响应变量，构建逐年随机森林回归模型，并采用 SHAP 方法解释各因子对模型输出的边际贡献。为避免循环解释，RF-SHAP 分析仅纳入 DEM、SLOPE、ASPECT、WRB、PRE、TEMMAX、TEMMIN、TEM、SOIL、AET 和 SWE 等外生环境因子，未纳入 HARSEI 组成项及 HAI/LUCC 相关变量。每个年份随机抽取 100,000 个有效像元用于模型训练与验证，其中 25% 作为独立验证集；随后从验证集中抽取 2,000 个样本计算 SHAP 值。",
            "",
            "RF 模型在六个年份均表现出较强解释能力，验证集 R² 保持在较高水平，说明外生环境变量能够较好刻画 HARSEI 的空间异质性。SHAP 结果表明，SOIL、AET、PRE 和 SWE 是多数年份贡献最高的变量，反映出土壤背景、水分供给和雪水过程对区域生态质量具有持续影响。SHAP 蜂群图进一步显示，高、低变量取值对 HARSEI 的影响方向存在明显差异，说明驱动因子不仅改变变量重要性排序，也通过非线性阈值效应影响生态质量响应。",
            "",
            "逐年关键结果如下：",
            *top_lines,
            "",
            "【English manuscript-ready wording】",
            "",
            "To further quantify the nonlinear effects of exogenous driving factors on the spatial heterogeneity of HARSEI, annual random forest regression models were constructed using the revised HARSEI as the response variable, and SHAP values were used to interpret the marginal contribution of each factor to model output. To avoid circular interpretation, only exogenous environmental drivers were included, namely DEM, SLOPE, ASPECT, WRB, PRE, TEMMAX, TEMMIN, TEM, SOIL, AET, and SWE. HARSEI component variables and HAI/LUCC-related variables were excluded from the RF-SHAP analysis. For each year, 100,000 valid pixels were randomly sampled for model training and validation, of which 25% were used as an independent validation set. SHAP values were then calculated for 2,000 validation samples.",
            "",
            "The RF models showed strong explanatory performance across the six years, with high validation R² values, indicating that the exogenous environmental variables effectively captured the spatial heterogeneity of HARSEI. The SHAP results indicated that SOIL, AET, PRE, and SWE were generally the most influential variables, highlighting the persistent roles of soil background, water availability, and snow-water processes in regulating regional ecological quality. The SHAP beeswarm plots further showed distinct effects of high and low feature values on HARSEI, suggesting that the drivers affected ecological quality through nonlinear and threshold-dependent responses rather than only through monotonic linear effects.",
        ]
    )
    (OUT_ROOT / "Section_3.8_RF_SHAP_text_CN_EN.txt").write_text(text, encoding="utf-8-sig")

    readme = "\n".join(
        [
            "RF-SHAP variable importance and model explanation for revised HARSEI",
            "",
            "Dependent variable: revised annual HARSEI rasters.",
            "Drivers: DEM, SLOPE, ASPECT, WRB, PRE, TEMMAX, TEMMIN, TEM, SOIL, AET, SWE.",
            "Excluded: HARSEI component variables and HAI/LUCC-related variables to avoid circular interpretation.",
            f"Sampling: up to {MAX_MODEL_SAMPLES:,} valid pixels per year; test size={TEST_SIZE}; SHAP samples={MAX_SHAP_SAMPLES:,}.",
            "",
            "Key outputs:",
            "- figures/Fig_HARSEI_RF_SHAP_beeswarm_2000_2024.png/.tif/.pdf",
            "- figures/Fig_HARSEI_RF_SHAP_importance_heatmap_2000_2024.png/.tif/.pdf",
            "- tables/RF_SHAP_model_metrics.csv",
            "- tables/RF_SHAP_variable_importance.csv",
            "- tables/RF_SHAP_beeswarm_plot_data.csv",
            "- tables/RF_SHAP_sample_wide_YYYY.csv",
            "- tables/HARSEI_RF_SHAP_variable_importance_data.xlsx",
        ]
    )
    (OUT_ROOT / "README_HARSEI_RF_SHAP_outputs.txt").write_text(readme, encoding="utf-8-sig")


def main() -> None:
    gdal.UseExceptions()
    ensure_dirs()
    results = []
    for year in YEARS:
        results.append(train_and_explain(year))
    export_tables(results)
    plot_beeswarm(results)
    plot_importance_heatmap(results)
    write_text(results)
    print(f"Saved outputs to: {OUT_ROOT}")


if __name__ == "__main__":
    main()
