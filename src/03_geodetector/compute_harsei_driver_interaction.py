#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Recompute GeoDetector interaction detection using revised annual HARSEI.

The script uses revised HARSEI rasters as the dependent variable and the
exogenous driver rasters stored in the original data folder as explanatory
variables. Continuous drivers are discretized by seven quantile classes,
following the previous GeoDetector workflow; WRB is retained as a categorical
soil-group factor.
"""

from __future__ import annotations

import csv
import math
import os
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.lines import Line2D
from osgeo import gdal

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover - workbook is a convenience export
    Workbook = None


YEARS = [2000, 2005, 2010, 2015, 2020, 2024]
K_CLASSES = 7

ROOT = Path(r"D:\Codex\260724 小论文")
HARSEI_DIR = ROOT / "revise" / "annual_harsei_outputs" / "rasters" / "HARSEI"
DRIVER_ROOT = Path(r"E:\wl24\hys to wl\数据清单\1数据下载")
OUT_ROOT = ROOT / "意见修改" / "图表" / "HARSEI_驱动因子交互探测"
FIG_DIR = OUT_ROOT / "figures"
TABLE_DIR = OUT_ROOT / "tables"
RUN_DIR = ROOT / "runs" / "20260811-geodetector-interaction"


DRIVERS: Dict[str, Dict[str, object]] = {
    "DEM": {
        "kind": "continuous",
        "path": DRIVER_ROOT / "24DEM" / "2_mask" / "masked_DEM_30m_shp.tif",
    },
    "SLOPE": {
        "kind": "continuous",
        "path": DRIVER_ROOT / "24DEM" / "2_mask" / "masked_Slope_30m_shp_deg.tif",
    },
    "ASPECT": {
        "kind": "continuous",
        "path": DRIVER_ROOT / "24DEM" / "2_mask" / "masked_Aspect_30m_shp_deg.tif",
    },
    "WRB": {
        "kind": "categorical",
        "path": DRIVER_ROOT / "30WRB" / "WRB4.tif",
    },
    "PRE": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "23Pre" / "2_mask" / "masked_pr_{year}.tif",
    },
    "TEMMAX": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "22Temmax" / "2_mask" / "masked_tem_{year}.tif",
    },
    "TEMMIN": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "25TEMmin" / "{year}.tif",
    },
    "TEM": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "26TEMave" / "{year}.tif",
    },
    "SOIL": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "27soil" / "soil{year}.tif",
    },
    "AET": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "28AET" / "aet{year}.tif",
    },
    "SWE": {
        "kind": "continuous",
        "template": DRIVER_ROOT / "29swe" / "swe{year}.tif",
    },
}

VAR_ORDER = ["DEM", "SLOPE", "ASPECT", "WRB", "PRE", "TEMMAX", "TEMMIN", "TEM", "SOIL", "AET", "SWE"]


def ensure_dirs() -> None:
    for folder in (FIG_DIR, TABLE_DIR, RUN_DIR):
        folder.mkdir(parents=True, exist_ok=True)


def driver_path(name: str, year: int) -> Path:
    meta = DRIVERS[name]
    if "path" in meta:
        return Path(meta["path"])
    return Path(str(meta["template"]).format(year=year))


def read_raster(path: Path, ref_ds=None, resample: str = "bilinear") -> np.ndarray:
    if not path.exists():
        raise FileNotFoundError(str(path))
    ds = gdal.Open(str(path))
    if ds is None:
        raise RuntimeError(f"Cannot open raster: {path}")

    if ref_ds is not None:
        same_grid = (
            ds.RasterXSize == ref_ds.RasterXSize
            and ds.RasterYSize == ref_ds.RasterYSize
            and tuple(round(v, 12) for v in ds.GetGeoTransform()) == tuple(round(v, 12) for v in ref_ds.GetGeoTransform())
            and ds.GetProjection() == ref_ds.GetProjection()
        )
        if not same_grid:
            gt = ref_ds.GetGeoTransform()
            xmin = gt[0]
            ymax = gt[3]
            xmax = xmin + gt[1] * ref_ds.RasterXSize
            ymin = ymax + gt[5] * ref_ds.RasterYSize
            alg = gdal.GRA_NearestNeighbour if resample == "nearest" else gdal.GRA_Bilinear
            ds = gdal.Warp(
                "",
                ds,
                format="MEM",
                dstSRS=ref_ds.GetProjection(),
                outputBounds=(xmin, ymin, xmax, ymax),
                width=ref_ds.RasterXSize,
                height=ref_ds.RasterYSize,
                resampleAlg=alg,
            )

    band = ds.GetRasterBand(1)
    arr = band.ReadAsArray().astype(np.float64)
    nodata = band.GetNoDataValue()
    if nodata is not None:
        arr[np.isclose(arr, nodata)] = np.nan
    arr[~np.isfinite(arr)] = np.nan
    # Common GDAL float nodata sentinels are sometimes not exactly registered.
    arr[arr < -1.0e20] = np.nan
    arr[arr > 1.0e20] = np.nan
    return arr


def quantile_classes(values: np.ndarray, k: int = K_CLASSES) -> Tuple[np.ndarray, List[float]]:
    classes = np.zeros(values.shape, dtype=np.uint16)
    valid = np.isfinite(values)
    vals = values[valid]
    if vals.size == 0:
        return classes, []
    edges = np.nanquantile(vals, np.linspace(0.0, 1.0, k + 1))
    edges[0] = -np.inf
    edges[-1] = np.inf
    # If duplicate quantile edges occur, digitize still works but may skip a few
    # empty class ids; GeoDetector only requires stable categorical strata.
    classes[valid] = np.digitize(values[valid], edges[1:-1], right=True).astype(np.uint16) + 1
    return classes, [float(x) if np.isfinite(x) else x for x in edges]


def categorical_classes(values: np.ndarray) -> Tuple[np.ndarray, List[float]]:
    classes = np.zeros(values.shape, dtype=np.uint16)
    valid = np.isfinite(values)
    vals = np.rint(values[valid]).astype(np.int64)
    uniques = sorted(int(v) for v in np.unique(vals))
    mapping = {value: i + 1 for i, value in enumerate(uniques)}
    out = np.zeros(vals.shape, dtype=np.uint16)
    for value, cls in mapping.items():
        out[vals == value] = cls
    classes[valid] = out
    return classes, [float(v) for v in uniques]


def q_stat(y: np.ndarray, strata: np.ndarray) -> float:
    mask = np.isfinite(y) & (strata > 0)
    yy = y[mask]
    ss = strata[mask]
    n = yy.size
    if n < 2:
        return float("nan")
    total_var = float(np.var(yy, ddof=0))
    if total_var <= 0:
        return float("nan")
    within = 0.0
    for cls in np.unique(ss):
        group = yy[ss == cls]
        if group.size:
            within += group.size * float(np.var(group, ddof=0))
    q = 1.0 - within / (n * total_var)
    return max(0.0, min(1.0, float(q)))


def overlay_strata(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    out = np.zeros(a.shape, dtype=np.uint32)
    mask = (a > 0) & (b > 0)
    max_b = int(np.nanmax(b)) if np.any(b > 0) else 0
    out[mask] = (a[mask].astype(np.uint32) - 1) * (max_b + 1) + b[mask].astype(np.uint32)
    return out


def interaction_type(q1: float, q2: float, q12: float) -> str:
    if any(math.isnan(x) for x in (q1, q2, q12)):
        return "NA"
    mn = min(q1, q2)
    mx = max(q1, q2)
    sm = q1 + q2
    eps = 1.0e-6
    if q12 < mn - eps:
        return "nonlinear weakening"
    if mn - eps <= q12 < mx - eps:
        return "single-factor weakening"
    if abs(q12 - sm) <= eps:
        return "independent"
    if mx + eps <= q12 < sm - eps:
        return "bivariate enhancement"
    if q12 >= sm + eps:
        return "nonlinear enhancement"
    return "independent"


def write_matrix_csv(path: Path, labels: List[str], matrix: np.ndarray, fmt="{:.10f}") -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([""] + labels)
        for label, row in zip(labels, matrix):
            writer.writerow([label] + [fmt.format(float(x)) if np.isfinite(float(x)) else "" for x in row])


def write_type_csv(path: Path, labels: List[str], matrix: List[List[str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([""] + labels)
        for label, row in zip(labels, matrix):
            writer.writerow([label] + row)


def compute_year(year: int) -> Tuple[Dict[str, float], np.ndarray, List[List[str]], Dict[str, List[float]], int]:
    harsei_path = HARSEI_DIR / f"YJQ_HARSEI_{year}.tif"
    ref_ds = gdal.Open(str(harsei_path))
    if ref_ds is None:
        raise RuntimeError(f"Cannot open HARSEI raster: {harsei_path}")
    y = read_raster(harsei_path)
    y[(y < -1000) | (y > 1000)] = np.nan

    strata: Dict[str, np.ndarray] = {}
    class_edges: Dict[str, List[float]] = {}
    for name in VAR_ORDER:
        meta = DRIVERS[name]
        resample = "nearest" if meta["kind"] == "categorical" else "bilinear"
        arr = read_raster(driver_path(name, year), ref_ds=ref_ds, resample=resample)
        if meta["kind"] == "categorical":
            cls, edges = categorical_classes(arr)
        else:
            cls, edges = quantile_classes(arr)
        strata[name] = cls
        class_edges[name] = edges

    common = np.isfinite(y)
    for name in VAR_ORDER:
        common &= strata[name] > 0
    n_common = int(common.sum())
    yy = np.where(common, y, np.nan)
    for name in VAR_ORDER:
        strata[name] = np.where(common, strata[name], 0)

    single_q = {name: q_stat(yy, strata[name]) for name in VAR_ORDER}
    nvar = len(VAR_ORDER)
    interaction_q = np.zeros((nvar, nvar), dtype=np.float64)
    type_matrix: List[List[str]] = [["single" for _ in VAR_ORDER] for _ in VAR_ORDER]
    for i, name_i in enumerate(VAR_ORDER):
        for j, name_j in enumerate(VAR_ORDER):
            if i == j:
                interaction_q[i, j] = single_q[name_i]
                type_matrix[i][j] = "single"
            elif j < i:
                interaction_q[i, j] = interaction_q[j, i]
                type_matrix[i][j] = type_matrix[j][i]
            else:
                combo = overlay_strata(strata[name_i], strata[name_j])
                q12 = q_stat(yy, combo)
                interaction_q[i, j] = q12
                interaction_q[j, i] = q12
                kind = interaction_type(single_q[name_i], single_q[name_j], q12)
                type_matrix[i][j] = kind
                type_matrix[j][i] = kind
    return single_q, interaction_q, type_matrix, class_edges, n_common


def plot_panel(all_q: Dict[int, np.ndarray], all_types: Dict[int, List[List[str]]], single_qs: Dict[int, Dict[str, float]]) -> None:
    labels = VAR_ORDER
    n = len(labels)
    vmax = max(float(np.nanmax(q)) for q in all_q.values())
    cmap = LinearSegmentedColormap.from_list(
        "deep_teal_gold",
        ["#16213e", "#29366f", "#315f86", "#2f9c95", "#93d66b", "#ffe26a"],
        N=256,
    )

    fig, axes = plt.subplots(2, 3, figsize=(15.2, 10.2), constrained_layout=False)
    fig.subplots_adjust(left=0.07, right=0.93, top=0.92, bottom=0.11, wspace=0.23, hspace=0.30)
    letters = ["a", "b", "c", "d", "e", "f"]
    last_im = None
    for ax, year, letter in zip(axes.flat, YEARS, letters):
        mat = all_q[year]
        last_im = ax.imshow(mat, cmap=cmap, vmin=0.0, vmax=vmax, interpolation="nearest")
        ax.set_title(f"({letter}) {year}", loc="left", fontsize=12, fontweight="bold")
        ax.set_xticks(range(n))
        ax.set_yticks(range(n))
        ax.set_xticklabels(labels, rotation=50, ha="right", fontsize=8)
        ax.set_yticklabels(labels, fontsize=8)
        ax.tick_params(length=0)
        ax.set_xlim(-0.5, n - 0.5)
        ax.set_ylim(n - 0.5, -0.5)
        ax.set_aspect("equal")

        for pos in np.arange(-0.5, n, 1.0):
            ax.axhline(pos, color="#f7f7f7", lw=0.45, alpha=0.75)
            ax.axvline(pos, color="#f7f7f7", lw=0.45, alpha=0.75)

        types = all_types[year]
        for i in range(n):
            for j in range(i + 1, n):
                kind = types[i][j]
                if kind == "nonlinear enhancement":
                    ax.scatter(j, i, marker="o", s=19, facecolor="#2db27d", edgecolor="white", linewidth=0.35, zorder=3)
                elif kind == "bivariate enhancement":
                    ax.scatter(j, i, marker="^", s=22, facecolor="#ff6b57", edgecolor="white", linewidth=0.35, zorder=3)
                elif kind == "independent":
                    ax.scatter(j, i, marker="s", s=13, facecolor="#f6f1d1", edgecolor="#555555", linewidth=0.25, zorder=3)
                elif "weakening" in kind:
                    ax.scatter(j, i, marker="x", s=21, color="#d0d0d0", linewidth=0.9, zorder=3)

        # Put single-factor q values on the diagonal for readability.
        for i, name in enumerate(labels):
            ax.text(i, i, f"{single_qs[year][name]:.2f}", ha="center", va="center", fontsize=6.5, color="white", fontweight="bold")

    cax = fig.add_axes([0.945, 0.18, 0.015, 0.66])
    cb = fig.colorbar(last_im, cax=cax)
    cb.set_label("Interaction q statistic", fontsize=10)
    cb.ax.tick_params(labelsize=8)

    handles = [
        Line2D([0], [0], marker="^", color="none", markerfacecolor="#ff6b57", markeredgecolor="white", markersize=7, label="Bivariate enhancement"),
        Line2D([0], [0], marker="o", color="none", markerfacecolor="#2db27d", markeredgecolor="white", markersize=7, label="Nonlinear enhancement"),
    ]
    fig.legend(handles=handles, loc="lower center", ncol=2, frameon=False, fontsize=9)
    fig.suptitle("GeoDetector interaction detection for revised HARSEI", fontsize=15, fontweight="bold")

    for ext, dpi in [("png", 450), ("tif", 450), ("pdf", 450)]:
        out = FIG_DIR / f"Fig_HARSEI_GeoDetector_interaction_2000_2024.{ext}"
        if ext == "tif":
            fig.savefig(out, dpi=dpi, pil_kwargs={"compression": "tiff_lzw"})
        else:
            fig.savefig(out, dpi=dpi)
    plt.close(fig)


def write_workbook(
    single_qs: Dict[int, Dict[str, float]],
    all_q: Dict[int, np.ndarray],
    all_types: Dict[int, List[List[str]]],
    class_edges: Dict[int, Dict[str, List[float]]],
    n_common: Dict[int, int],
) -> None:
    if Workbook is None:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "single_factor_q"
    ws.append(["Year", "N_common", "Factor", "q"])
    for year in YEARS:
        for name in VAR_ORDER:
            ws.append([year, n_common[year], name, single_qs[year][name]])

    ws = wb.create_sheet("interaction_q_long")
    ws.append(["Year", "Factor_1", "Factor_2", "q_1", "q_2", "q_12", "Interaction_type"])
    for year in YEARS:
        mat = all_q[year]
        typ = all_types[year]
        for i, a in enumerate(VAR_ORDER):
            for j, b in enumerate(VAR_ORDER):
                if j <= i:
                    continue
                ws.append([year, a, b, single_qs[year][a], single_qs[year][b], float(mat[i, j]), typ[i][j]])

    for year in YEARS:
        ws = wb.create_sheet(f"q_matrix_{year}")
        ws.append([""] + VAR_ORDER)
        for label, row in zip(VAR_ORDER, all_q[year]):
            ws.append([label] + [float(v) for v in row])
        ws = wb.create_sheet(f"type_matrix_{year}")
        ws.append([""] + VAR_ORDER)
        for label, row in zip(VAR_ORDER, all_types[year]):
            ws.append([label] + row)

    ws = wb.create_sheet("strata_edges")
    ws.append(["Year", "Factor", "Classing", "Edges_or_categories"])
    for year in YEARS:
        for name in VAR_ORDER:
            kind = "categorical" if DRIVERS[name]["kind"] == "categorical" else f"{K_CLASSES}-class quantile"
            ws.append([year, name, kind, ", ".join(str(x) for x in class_edges[year][name])])

    wb.save(TABLE_DIR / "HARSEI_GeoDetector_interaction_data.xlsx")


def main() -> None:
    gdal.UseExceptions()
    ensure_dirs()
    single_qs: Dict[int, Dict[str, float]] = {}
    all_q: Dict[int, np.ndarray] = {}
    all_types: Dict[int, List[List[str]]] = {}
    all_edges: Dict[int, Dict[str, List[float]]] = {}
    n_common: Dict[int, int] = {}

    long_rows: List[List[object]] = []
    for year in YEARS:
        print(f"Processing {year}...")
        single_q, interaction_q, type_matrix, edges, n = compute_year(year)
        single_qs[year] = single_q
        all_q[year] = interaction_q
        all_types[year] = type_matrix
        all_edges[year] = edges
        n_common[year] = n

        with (TABLE_DIR / f"GD_factor_revised_HARSEI_{year}.csv").open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Year", "N_common", "Factor", "q", "classing"])
            for name in VAR_ORDER:
                kind = "categorical" if DRIVERS[name]["kind"] == "categorical" else f"{K_CLASSES}-class quantile"
                writer.writerow([year, n, name, f"{single_q[name]:.10f}", kind])
        write_matrix_csv(TABLE_DIR / f"GD_interaction_q_revised_HARSEI_{year}.csv", VAR_ORDER, interaction_q)
        write_type_csv(TABLE_DIR / f"GD_interaction_type_revised_HARSEI_{year}.csv", VAR_ORDER, type_matrix)

        for i, a in enumerate(VAR_ORDER):
            for j, b in enumerate(VAR_ORDER):
                if j <= i:
                    continue
                long_rows.append([year, a, b, single_q[a], single_q[b], interaction_q[i, j], type_matrix[i][j]])

    with (TABLE_DIR / "GD_interaction_long_revised_HARSEI_2000_2024.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Year", "Factor_1", "Factor_2", "q_1", "q_2", "q_12", "Interaction_type"])
        writer.writerows(long_rows)

    with (TABLE_DIR / "GD_common_valid_pixels_revised_HARSEI.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Year", "N_common"])
        for year in YEARS:
            writer.writerow([year, n_common[year]])

    write_workbook(single_qs, all_q, all_types, all_edges, n_common)
    plot_panel(all_q, all_types, single_qs)

    readme = OUT_ROOT / "README_HARSEI_GeoDetector_interaction.txt"
    readme.write_text(
        "\n".join(
            [
                "GeoDetector interaction detection for revised annual HARSEI",
                "",
                "Dependent variable: revised HARSEI rasters from D:\\Codex\\260724 小论文\\revise\\annual_harsei_outputs\\rasters\\HARSEI.",
                "Driver variables: DEM, SLOPE, ASPECT, WRB, PRE, TEMMAX, TEMMIN, TEM, SOIL, AET, and SWE from E:\\wl24\\hys to wl\\数据清单\\1数据下载.",
                "Excluded from the main driver-interaction figure: HARSEI component variables and HAI/LUCC-related variables, to avoid circularity in driver interpretation.",
                f"Continuous drivers were discretized into {K_CLASSES} quantile classes for each year; WRB was retained as a categorical factor.",
                "Interaction type follows standard GeoDetector rules based on q(X1), q(X2), q(X1∩X2), and q(X1)+q(X2).",
                "",
                "Key outputs:",
                f"- figures/Fig_HARSEI_GeoDetector_interaction_2000_2024.png/.tif/.pdf",
                "- tables/GD_factor_revised_HARSEI_YYYY.csv",
                "- tables/GD_interaction_q_revised_HARSEI_YYYY.csv",
                "- tables/GD_interaction_type_revised_HARSEI_YYYY.csv",
                "- tables/GD_interaction_long_revised_HARSEI_2000_2024.csv",
                "- tables/HARSEI_GeoDetector_interaction_data.xlsx",
            ]
        ),
        encoding="utf-8-sig",
    )

    print(f"Saved outputs to: {OUT_ROOT}")


if __name__ == "__main__":
    main()
